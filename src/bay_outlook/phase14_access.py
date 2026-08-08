from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import shutil
import sqlite3
import subprocess
import tempfile
import time
import urllib.request
import zipfile
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .constants import COUNTY_BY_FIPS, FIPS_BY_COUNTY, PROJECT_ROOT
from .phase14 import _number, _ratio, _sum, verify_phase14
from .storage import USER_AGENT, save_snapshot, sha256_bytes


PHASE14_ACCESS_VERSION = "1.2.0"
ACS_YEARS = (2021, 2022, 2023, 2024)
PIT_HIC_YEARS = (2022, 2023, 2024, 2025)
PICTURE_YEARS = (2022, 2023, 2024, 2025)

REQUIRED_ACCESS_DOMAINS = {
    "displacement_pressure",
    "eviction_filings",
    "homelessness",
    "hud_assisted_housing",
    "income_targeted_production",
    "worker_housing_access",
}

CRITICAL_ACCESS_METRICS = {
    "displacement_pressure": "renter_overcrowding_pct",
    "eviction_filings": "unlawful_detainer_filings",
    "homelessness": "pit_overall_homeless",
    "hud_assisted_housing": "hud_assisted_units",
    "income_targeted_production": "hcd_apr_vli_permitted_units",
    "worker_housing_access": "median_worker_earnings",
}

COUNTY_TO_COC = {
    "06001": "CA-502",
    "06013": "CA-505",
    "06041": "CA-507",
    "06055": "CA-517",
    "06075": "CA-501",
    "06081": "CA-512",
    "06085": "CA-500",
    "06095": "CA-518",
    "06097": "CA-504",
}

ACCESS_SOURCE_REGISTRY = {
    "CA_COURTS_COURT_STATISTICS": {
        "publisher": "Judicial Council of California",
        "title": "Court Statistics Report — Filings and Dispositions",
        "source_tier": 1,
        "source_class": "government administrative data",
        "frequency": "annual fiscal year",
        "geography_basis": "county superior court",
        "landing_url": "https://courts.ca.gov/policy-administration/research-analytics/data-statistics",
    },
    "CENSUS_ACS5_BULK": {
        "publisher": "U.S. Census Bureau",
        "title": "American Community Survey 5-Year Detailed Tables",
        "source_tier": 1,
        "source_class": "government survey estimate",
        "frequency": "annual five-year estimate",
        "geography_basis": "household residence",
        "landing_url": "https://www.census.gov/programs-surveys/acs/data.html",
    },
    "HUD_PIT": {
        "publisher": "U.S. Department of Housing and Urban Development",
        "title": "Point-in-Time Estimates by Continuum of Care",
        "source_tier": 1,
        "source_class": "government administrative count",
        "frequency": "annual January count",
        "geography_basis": "Continuum of Care mapped one-to-one to Bay Area county",
        "landing_url": "https://www.huduser.gov/portal/datasets/ahar/2025-ahar-part-1-pit-estimates-of-homelessness-in-the-us.html",
    },
    "HUD_HIC": {
        "publisher": "U.S. Department of Housing and Urban Development",
        "title": "Housing Inventory Count by Continuum of Care",
        "source_tier": 1,
        "source_class": "government administrative inventory",
        "frequency": "annual January inventory",
        "geography_basis": "Continuum of Care mapped one-to-one to Bay Area county",
        "landing_url": "https://www.huduser.gov/portal/datasets/ahar/2025-ahar-part-1-pit-estimates-of-homelessness-in-the-us.html",
    },
    "HUD_PICTURE": {
        "publisher": "U.S. Department of Housing and Urban Development",
        "title": "Picture of Subsidized Households — County Summary",
        "source_tier": 1,
        "source_class": "government administrative inventory",
        "frequency": "annual",
        "geography_basis": "county program administration",
        "landing_url": "https://www.huduser.gov/portal/datasets/assthsg.html",
    },
    "HUD_FMR": {
        "publisher": "U.S. Department of Housing and Urban Development",
        "title": "FY 2026 Fair Market Rents, Revised",
        "source_tier": 1,
        "source_class": "government modeled rent standard",
        "frequency": "federal fiscal year",
        "geography_basis": "HUD FMR area assigned to county",
        "landing_url": "https://www.huduser.gov/portal/datasets/fmr.html",
    },
    "CA_HCD_APR": {
        "publisher": "California Department of Housing and Community Development",
        "title": "Housing Element Annual Progress Report — Table A2",
        "source_tier": 1,
        "source_class": "government administrative data",
        "frequency": "annual with portal updates",
        "geography_basis": "reporting jurisdiction aggregated to county",
        "landing_url": "https://www.hcd.ca.gov/housing-open-data-tools/apr-dashboard",
    },
    "CA_HCD_RHNA": {
        "publisher": "California Department of Housing and Community Development",
        "title": "Sixth-Cycle Regional Housing Needs Allocation",
        "source_tier": 1,
        "source_class": "government planning allocation",
        "frequency": "planning cycle",
        "geography_basis": "jurisdiction allocation aggregated to county",
        "landing_url": "https://www.hcd.ca.gov/rhna",
    },
}

COURT_URL = "https://courts.ca.gov/system/files/file/filings-and-dispositions.csv"
HUD_PIT_URL = "https://www.huduser.gov/portal/sites/default/files/xls/2007-2025-PIT-Counts-by-CoC.xlsb"
HUD_HIC_URL = "https://www.huduser.gov/portal/sites/default/files/xls/2007-2025-HIC-Counts-by-CoC.xlsx"
HUD_PICTURE_URL = (
    "https://www.huduser.gov/portal/datasets/pictures/files/"
    "COUNTY_{year}_{suffix}.xlsx"
)
HUD_PICTURE_SUFFIX = {2022: "2020census", 2023: "2020census", 2024: "2020census", 2025: "2020census"}
HUD_FMR_URL = "https://www.huduser.gov/portal/datasets/fmr/fmr2026/FY26_FMRs_revised.xlsx"
ACS_BASE = (
    "https://www2.census.gov/programs-surveys/acs/summary_file/"
    "{year}/table-based-SF/data/5YRData/acsdt5y{year}-{table}.dat"
)

OBSERVATION_FIELDS = (
    "domain",
    "metric_id",
    "metric_name",
    "county_fips",
    "county_name",
    "period",
    "period_end",
    "frequency",
    "value",
    "unit",
    "margin_of_error",
    "source_ids",
    "source_releases",
    "retrieved_at",
    "raw_sha256s",
    "calculation",
    "notes",
    "comparability_status",
)

REQUIRED_ACCESS_FILES = (
    "config/phase14_housing_access.json",
    "metadata/housing_access_indicator_catalog.csv",
    "metadata/housing_access_source_registry.csv",
    "docs/housing-access/README.md",
    "docs/housing-access/METHODOLOGY.md",
    "docs/housing-access/DATA_DICTIONARY.md",
    "docs/housing-access/LIMITATIONS.md",
    "docs/housing-access/RUNBOOK.md",
    "src/bay_outlook/phase14_access.py",
    "tests/test_phase14_access.py",
    ".github/workflows/housing-access-update.yml",
    "data/phase14/access/exports/access_snapshot.csv",
    "data/phase14/access/exports/income_progress.csv",
    "data/phase14/access/exports/pit_count_types.csv",
    "data/phase14/access/exports/source_freshness.csv",
    "data/phase14/access/exports/quality_checks.csv",
)


def _utc_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: Iterable[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    names = list(fieldnames or (rows[0].keys() if rows else []))
    with path.open("w", newline="", encoding="utf-8") as stream:
        writer = csv.DictWriter(stream, fieldnames=names, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _observation(
    *,
    domain: str,
    metric_id: str,
    metric_name: str,
    county_fips: str,
    period: str,
    period_end: str,
    frequency: str,
    value: float | None,
    unit: str,
    source_ids: str,
    source_releases: str,
    retrieved_at: str,
    raw_sha256s: str,
    margin_of_error: float | None = None,
    calculation: str = "published",
    notes: str = "",
    comparability_status: str = "comparable_within_series",
) -> dict[str, Any]:
    return {
        "domain": domain,
        "metric_id": metric_id,
        "metric_name": metric_name,
        "county_fips": county_fips,
        "county_name": COUNTY_BY_FIPS[county_fips],
        "period": period,
        "period_end": period_end,
        "frequency": frequency,
        "value": value,
        "unit": unit,
        "margin_of_error": margin_of_error,
        "source_ids": source_ids,
        "source_releases": source_releases,
        "retrieved_at": retrieved_at,
        "raw_sha256s": raw_sha256s,
        "calculation": calculation,
        "notes": notes,
        "comparability_status": comparability_status,
    }


def _request_with_retries(
    url: str,
    *,
    landing_url: str | None = None,
    timeout: int = 240,
    attempts: int = 5,
) -> tuple[bytes, dict[str, str], str]:
    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "application/octet-stream,text/csv,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
    }
    if landing_url:
        headers["Referer"] = landing_url
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            request = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(request, timeout=timeout) as response:
                content = response.read()
                response_headers = dict(response.headers.items())
                final_url = response.geturl()
            if len(content) < 100:
                raise ValueError(f"download returned only {len(content)} bytes")
            return content, response_headers, final_url
        except Exception as exc:  # pragma: no cover - only exercised on transient upstream faults
            last_error = exc
            if landing_url:
                try:
                    warm = urllib.request.Request(landing_url, headers={"User-Agent": USER_AGENT})
                    with urllib.request.urlopen(warm, timeout=60) as response:
                        response.read(1024)
                except Exception:
                    pass
            if attempt < attempts:
                time.sleep(min(2 * attempt, 8))
    raise RuntimeError(f"unable to retrieve {url}: {last_error}")


def _snapshot_record(snapshot: Any, metadata: dict[str, Any], root: Path) -> dict[str, Any]:
    return {
        "source_id": snapshot.source_id,
        "source_url": snapshot.source_url,
        "source_release": snapshot.source_release,
        "retrieved_at": snapshot.retrieved_at,
        "snapshot_sha256": snapshot.sha256,
        "upstream_sha256": metadata.get("upstream_sha256", snapshot.sha256),
        "byte_count": metadata.get("byte_count"),
        "upstream_byte_count": metadata.get("upstream_byte_count", metadata.get("byte_count")),
        "snapshot_kind": metadata.get("snapshot_kind", "complete_response"),
        "raw_path": str(snapshot.path.relative_to(root)),
        "metadata_path": str(snapshot.metadata_path.relative_to(root)),
    }


def _save_complete_snapshot(
    *,
    root: Path,
    data_root: Path,
    source_id: str,
    filename: str,
    url: str,
    release: str,
    retrieved_at: str,
    landing_url: str | None = None,
) -> tuple[bytes, dict[str, Any], Path]:
    same_day = data_root / "raw" / source_id / retrieved_at[:10] / filename
    same_day_metadata = Path(f"{same_day}.metadata.json")
    if same_day.is_file() and same_day_metadata.is_file():
        metadata = _read_json(same_day_metadata)
        content = same_day.read_bytes()
        if sha256_bytes(content) != metadata.get("sha256"):
            raise ValueError(f"same-day snapshot hash mismatch: {same_day}")
        snapshot = type(
            "ArchivedSnapshot",
            (),
            {
                "source_id": source_id,
                "source_url": metadata["source_url"],
                "source_release": metadata["source_release"],
                "retrieved_at": metadata["retrieved_at"],
                "sha256": metadata["sha256"],
                "path": same_day,
                "metadata_path": same_day_metadata,
            },
        )()
        return content, _snapshot_record(snapshot, metadata, root), same_day
    try:
        content, headers, final_url = _request_with_retries(url, landing_url=landing_url)
    except RuntimeError:
        archived = sorted((data_root / "raw" / source_id).glob(f"*/{filename}"), reverse=True)
        if not archived:
            raise
        archived_path = archived[0]
        metadata_path = Path(f"{archived_path}.metadata.json")
        if not metadata_path.is_file():
            raise
        metadata = _read_json(metadata_path)
        content = archived_path.read_bytes()
        if sha256_bytes(content) != metadata.get("sha256"):
            raise ValueError(f"archived fallback hash mismatch: {archived_path}")
        metadata["last_refresh_fallback_at"] = retrieved_at
        metadata["last_refresh_fallback_reason"] = "upstream returned no usable content after retries"
        _write_json(metadata_path, metadata)
        snapshot = type(
            "ArchivedSnapshot",
            (),
            {
                "source_id": source_id,
                "source_url": metadata["source_url"],
                "source_release": metadata["source_release"],
                "retrieved_at": metadata["retrieved_at"],
                "sha256": metadata["sha256"],
                "path": archived_path,
                "metadata_path": metadata_path,
            },
        )()
        return content, _snapshot_record(snapshot, metadata, root), archived_path
    snapshot = save_snapshot(
        data_root,
        source_id,
        filename,
        content,
        final_url,
        release,
        response_headers=headers,
        retrieved_at=retrieved_at,
    )
    metadata = _read_json(snapshot.metadata_path)
    return content, _snapshot_record(snapshot, metadata, root), snapshot.path


def _fetch_acs_extract(
    *,
    root: Path,
    data_root: Path,
    year: int,
    table: str,
    retrieved_at: str,
) -> tuple[dict[str, dict[str, str]], dict[str, Any]]:
    url = ACS_BASE.format(year=year, table=table.casefold())
    filename = f"acs5-{year}-{table.casefold()}-bay-counties.dat"
    same_day = data_root / "raw" / "CENSUS_ACS5_BULK" / retrieved_at[:10] / filename
    same_day_metadata = Path(f"{same_day}.metadata.json")
    if same_day.is_file() and same_day_metadata.is_file():
        content = same_day.read_bytes()
        metadata = _read_json(same_day_metadata)
        if sha256_bytes(content) != metadata.get("sha256"):
            raise ValueError(f"same-day ACS snapshot hash mismatch: {same_day}")
        snapshot = type(
            "ArchivedSnapshot",
            (),
            {
                "source_id": "CENSUS_ACS5_BULK",
                "source_url": metadata["source_url"],
                "source_release": metadata["source_release"],
                "retrieved_at": metadata["retrieved_at"],
                "sha256": metadata["sha256"],
                "path": same_day,
                "metadata_path": same_day_metadata,
            },
        )()
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")), delimiter="|")
        rows = {row["GEO_ID"][-5:]: row for row in reader}
        return rows, _snapshot_record(snapshot, metadata, root)
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    digest = hashlib.sha256()
    byte_count = 0
    selected: list[bytes] = []
    wanted = {f"0500000US{fips}".encode("ascii") for fips in COUNTY_BY_FIPS}
    headers: dict[str, str] = {}
    final_url = url
    with urllib.request.urlopen(request, timeout=240) as response:
        headers = dict(response.headers.items())
        final_url = response.geturl()
        for index, line in enumerate(response):
            digest.update(line)
            byte_count += len(line)
            if index == 0 or line.split(b"|", 1)[0] in wanted:
                selected.append(line)
    if len(selected) != len(COUNTY_BY_FIPS) + 1:
        raise ValueError(f"ACS {year} {table} extract did not contain all nine counties")
    content = b"".join(selected)
    snapshot = save_snapshot(
        data_root,
        "CENSUS_ACS5_BULK",
        filename,
        content,
        final_url,
        f"ACS5-{year}-{table}",
        response_headers=headers,
        retrieved_at=retrieved_at,
    )
    metadata = _read_json(snapshot.metadata_path)
    metadata.update(
        {
            "snapshot_kind": "source_slice",
            "upstream_sha256": digest.hexdigest(),
            "upstream_byte_count": byte_count,
            "selection": {
                "geography": "summary level 050 county rows for the nine configured Bay Area FIPS codes",
                "preserved_header": True,
                "row_count": len(COUNTY_BY_FIPS),
            },
        }
    )
    _write_json(snapshot.metadata_path, metadata)
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")), delimiter="|")
    rows = {row["GEO_ID"][-5:]: row for row in reader}
    return rows, _snapshot_record(snapshot, metadata, root)


def _normalise_count_type(value: str) -> str:
    return " ".join(value.replace("*", "").split())


def _looks_like_zip(content: bytes) -> bool:
    return content.startswith(b"PK\x03\x04")


def _repair_xlsx_properties(content: bytes) -> bytes:
    """Repair a malformed HUD core-property timestamp without changing sheet data."""
    if not _looks_like_zip(content):
        raise ValueError("expected an XLSX ZIP container")
    source = io.BytesIO(content)
    target = io.BytesIO()
    with zipfile.ZipFile(source) as input_zip, zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as output_zip:
        for info in input_zip.infolist():
            data = input_zip.read(info.filename)
            if info.filename == "docProps/core.xml":
                text = data.decode("utf-8")
                text = re.sub(
                    r"(\d{4})-\s*(\d{1,2})-\s*(\d{1,2})T",
                    lambda match: (
                        f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}T"
                    ),
                    text,
                )
                data = text.encode("utf-8")
            output_zip.writestr(info, data)
    return target.getvalue()


def _acs_access_observations(
    *, root: Path, data_root: Path, retrieved_at: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[tuple[str, str], dict[str, Any]]]:
    observations: list[dict[str, Any]] = []
    snapshots: list[dict[str, Any]] = []
    renter_household_rows: dict[tuple[str, str], dict[str, Any]] = {}
    tables = ("B25003", "B25014", "B20002")
    for year in ACS_YEARS:
        extracts: dict[str, dict[str, dict[str, str]]] = {}
        records: dict[str, dict[str, Any]] = {}
        for table in tables:
            extracts[table], records[table] = _fetch_acs_extract(
                root=root,
                data_root=data_root,
                year=year,
                table=table,
                retrieved_at=retrieved_at,
            )
            snapshots.append(records[table])
        for fips in COUNTY_BY_FIPS:
            tenure = extracts["B25003"][fips]
            renter_households = _number(tenure.get("B25003_E003"))
            renter_households_moe = _number(tenure.get("B25003_M003"))
            crowding = extracts["B25014"][fips]
            renter_total = _number(crowding.get("B25014_E008"))
            renter_overcrowded = _sum(
                _number(crowding.get(f"B25014_E{index:03d}")) for index in (11, 12, 13)
            )
            renter_severe = _sum(
                _number(crowding.get(f"B25014_E{index:03d}")) for index in (12, 13)
            )
            earnings = extracts["B20002"][fips]
            median_earnings = _number(earnings.get("B20002_E001"))
            median_earnings_moe = _number(earnings.get("B20002_M001"))
            common = {
                "county_fips": fips,
                "period": str(year),
                "period_end": f"{year}-12-31",
                "frequency": "annual five-year estimate",
                "retrieved_at": retrieved_at,
            }
            renter_row = _observation(
                domain="eviction_filings",
                metric_id="renter_households",
                metric_name="Renter-occupied households",
                value=renter_households,
                unit="households",
                margin_of_error=renter_households_moe,
                source_ids="CENSUS_ACS5_BULK",
                source_releases=f"ACS5-{year}-B25003",
                raw_sha256s=records["B25003"]["snapshot_sha256"],
                notes="ACS five-year estimate used as an exposure denominator, not a count of leases.",
                **common,
            )
            observations.append(renter_row)
            renter_household_rows[(fips, str(year))] = renter_row
            observations.extend(
                [
                    _observation(
                        domain="displacement_pressure",
                        metric_id="renter_overcrowding_pct",
                        metric_name="Renter households with more than one occupant per room",
                        value=_ratio(renter_overcrowded, renter_total),
                        unit="percent",
                        source_ids="CENSUS_ACS5_BULK",
                        source_releases=f"ACS5-{year}-B25014",
                        raw_sha256s=records["B25014"]["snapshot_sha256"],
                        calculation="B25014 renter categories above 1.00 occupants per room divided by renter-occupied units",
                        notes="Transparent displacement-pressure component; no composite risk score is calculated.",
                        **common,
                    ),
                    _observation(
                        domain="displacement_pressure",
                        metric_id="renter_severe_overcrowding_pct",
                        metric_name="Renter households with more than 1.5 occupants per room",
                        value=_ratio(renter_severe, renter_total),
                        unit="percent",
                        source_ids="CENSUS_ACS5_BULK",
                        source_releases=f"ACS5-{year}-B25014",
                        raw_sha256s=records["B25014"]["snapshot_sha256"],
                        calculation="B25014 renter categories above 1.50 occupants per room divided by renter-occupied units",
                        notes="Transparent displacement-pressure component; no composite risk score is calculated.",
                        **common,
                    ),
                    _observation(
                        domain="worker_housing_access",
                        metric_id="median_worker_earnings",
                        metric_name="Median annual earnings for people age 16+ with earnings",
                        value=median_earnings,
                        unit="current dollars per year",
                        margin_of_error=median_earnings_moe,
                        source_ids="CENSUS_ACS5_BULK",
                        source_releases=f"ACS5-{year}-B20002",
                        raw_sha256s=records["B20002"]["snapshot_sha256"],
                        notes="County median for people age 16 and over with earnings; not occupation-specific and not household income.",
                        **common,
                    ),
                ]
            )
    return observations, snapshots, renter_household_rows


def _inherited_v1_1_observations(
    *, root: Path, retrieved_at: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, dict[str, Any]]]:
    observations_path = root / "data" / "phase14" / "exports" / "housing_time_series.csv"
    snapshots_path = root / "data" / "phase14" / "exports" / "source_snapshots.csv"
    if not observations_path.is_file() or not snapshots_path.is_file():
        raise FileNotFoundError("Version 1.2 requires the verified Version 1.1 housing exports")
    with observations_path.open(newline="", encoding="utf-8") as stream:
        inherited = list(csv.DictReader(stream))
    with snapshots_path.open(newline="", encoding="utf-8") as stream:
        all_snapshots = list(csv.DictReader(stream))
    selected_snapshots = [
        row
        for row in all_snapshots
        if row["source_id"] in {"CENSUS_ACS5_BULK", "CA_HCD_APR", "CA_HCD_RHNA"}
    ]
    latest_by_source: dict[str, dict[str, Any]] = {}
    for row in selected_snapshots:
        current = latest_by_source.get(row["source_id"])
        if current is None or row["retrieved_at"] > current["retrieved_at"]:
            latest_by_source[row["source_id"]] = row

    observations: list[dict[str, Any]] = []
    for row in inherited:
        metric_id = row["metric_id"]
        if metric_id in {"renter_severe_cost_burden_50_plus_pct", "median_gross_rent"}:
            observations.append(
                _observation(
                    domain="displacement_pressure",
                    metric_id=metric_id,
                    metric_name=row["metric_name"],
                    county_fips=row["county_fips"],
                    period=row["period"],
                    period_end=row["period_end"],
                    frequency=row["frequency"],
                    value=_number(row["value"]),
                    unit=row["unit"],
                    margin_of_error=_number(row["margin_of_error"]),
                    source_ids=row["source_id"],
                    source_releases=row["source_release"],
                    retrieved_at=row["retrieved_at"] or retrieved_at,
                    raw_sha256s=row["raw_sha256"],
                    calculation=row["calculation"],
                    notes=(row["notes"] + " Reused from the verified Version 1.1 evidence layer.").strip(),
                )
            )
        if metric_id.startswith("hcd_apr_") and metric_id.endswith("_permitted_units"):
            observations.append(
                _observation(
                    domain="income_targeted_production",
                    metric_id=metric_id,
                    metric_name=row["metric_name"],
                    county_fips=row["county_fips"],
                    period=row["period"],
                    period_end=row["period_end"],
                    frequency=row["frequency"],
                    value=_number(row["value"]),
                    unit=row["unit"],
                    source_ids=row["source_id"],
                    source_releases=row["source_release"],
                    retrieved_at=row["retrieved_at"] or retrieved_at,
                    raw_sha256s=row["raw_sha256"],
                    calculation=row["calculation"],
                    notes=(row["notes"] + " Reused from the verified Version 1.1 evidence layer.").strip(),
                )
            )
    return observations, selected_snapshots, latest_by_source


def _court_observations(
    *,
    root: Path,
    data_root: Path,
    retrieved_at: str,
    renter_households: dict[tuple[str, str], dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    content, snapshot, _ = _save_complete_snapshot(
        root=root,
        data_root=data_root,
        source_id="CA_COURTS_COURT_STATISTICS",
        filename="filings-and-dispositions.csv",
        url=COURT_URL,
        release="California-Court-Statistics-through-FY2023-24",
        retrieved_at=retrieved_at,
        landing_url=ACCESS_SOURCE_REGISTRY["CA_COURTS_COURT_STATISTICS"]["landing_url"],
    )
    observations: list[dict[str, Any]] = []
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    found: set[tuple[str, str, str]] = set()
    for row in reader:
        if row.get("casetype_variable") != "Unlawful Detainer":
            continue
        county = row.get("county_name", "")
        county_key = county.casefold()
        if county_key not in FIPS_BY_COUNTY:
            continue
        year = str(row["fiscal_year"])
        disposition = row["filing_disposition"]
        fips = FIPS_BY_COUNTY[county_key]
        period = f"FY{year}-{str(int(year) + 1)[-2:]}"
        period_end = f"{int(year) + 1}-06-30"
        value = _number(row["summary_statistic"])
        metric_id = (
            "unlawful_detainer_filings" if disposition == "Filing" else "unlawful_detainer_dispositions"
        )
        metric_name = (
            "Limited civil unlawful-detainer filings"
            if disposition == "Filing"
            else "Limited civil unlawful-detainer dispositions"
        )
        observations.append(
            _observation(
                domain="eviction_filings",
                metric_id=metric_id,
                metric_name=metric_name,
                county_fips=fips,
                period=period,
                period_end=period_end,
                frequency="annual fiscal year",
                value=value,
                unit="court cases",
                source_ids="CA_COURTS_COURT_STATISTICS",
                source_releases=snapshot["source_release"],
                retrieved_at=retrieved_at,
                raw_sha256s=snapshot["snapshot_sha256"],
                notes="A court filing or disposition is not an executed eviction. Unlimited civil cases and informal displacement are not captured.",
                comparability_status="pandemic_policy_and_backlog_context_required",
            )
        )
        found.add((fips, year, disposition))
        denominator = renter_households.get((fips, year))
        if disposition == "Filing" and denominator and value is not None:
            observations.append(
                _observation(
                    domain="eviction_filings",
                    metric_id="unlawful_detainer_filings_per_1000_renter_households",
                    metric_name="Limited civil unlawful-detainer filings per 1,000 renter households",
                    county_fips=fips,
                    period=period,
                    period_end=period_end,
                    frequency="annual fiscal year with ACS denominator",
                    value=_ratio(value, denominator["value"], multiplier=1000.0),
                    unit="filings per 1,000 renter households",
                    source_ids="CA_COURTS_COURT_STATISTICS;CENSUS_ACS5_BULK",
                    source_releases=f"{snapshot['source_release']};{denominator['source_releases']}",
                    retrieved_at=retrieved_at,
                    raw_sha256s=f"{snapshot['snapshot_sha256']};{denominator['raw_sha256s']}",
                    calculation="limited civil unlawful-detainer filings divided by same-vintage ACS renter households, multiplied by 1,000",
                    notes="The ACS denominator is a five-year estimate aligned to the fiscal-year start, so this is an approximate exposure rate.",
                    comparability_status="derived_with_overlapping_acs_denominator",
                )
            )
    expected = {
        (fips, str(year), disposition)
        for fips in COUNTY_BY_FIPS
        for year in range(2013, 2024)
        for disposition in ("Filing", "Disposition")
    }
    if found != expected:
        raise ValueError(f"court extract coverage mismatch: missing={sorted(expected - found)[:10]}")
    return observations, [snapshot]


def _income_target_observations(
    *,
    root: Path,
    retrieved_at: str,
    latest_snapshots: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    path = root / "data" / "phase14" / "exports" / "rhna_progress.csv"
    with path.open(newline="", encoding="utf-8") as stream:
        rows = list(csv.DictReader(stream))
    snapshot = latest_snapshots.get("CA_HCD_RHNA")
    if snapshot is None:
        raise ValueError("Version 1.1 RHNA source snapshot is unavailable")
    observations: list[dict[str, Any]] = []
    metrics = {
        "rhna_vli": ("rhna_target_vli_units", "Sixth-cycle RHNA very-low-income allocation"),
        "rhna_li": ("rhna_target_li_units", "Sixth-cycle RHNA low-income allocation"),
        "rhna_mod": ("rhna_target_mod_units", "Sixth-cycle RHNA moderate-income allocation"),
        "rhna_above_mod": (
            "rhna_target_above_mod_units",
            "Sixth-cycle RHNA above-moderate-income allocation",
        ),
    }
    for row in rows:
        fips = row["county_fips"]
        for column, (metric_id, metric_name) in metrics.items():
            observations.append(
                _observation(
                    domain="income_targeted_production",
                    metric_id=metric_id,
                    metric_name=metric_name,
                    county_fips=fips,
                    period="sixth cycle",
                    period_end="2031-01-31",
                    frequency="planning cycle",
                    value=_number(row[column]),
                    unit="housing units",
                    source_ids="CA_HCD_RHNA",
                    source_releases="HCD-RHNA-sixth-cycle",
                    retrieved_at=snapshot["retrieved_at"] or retrieved_at,
                    raw_sha256s=snapshot["snapshot_sha256"],
                    calculation="sum of jurisdiction allocations within county",
                    notes="Allocation is a planning target, not a completed or occupied unit count.",
                    comparability_status="planning_target_not_observed_supply",
                )
            )
    return observations


def _extract_pit_sheets(
    raw_path: Path,
    years: Iterable[int],
    *,
    cache_dir: Path,
) -> dict[int, list[dict[str, str]]]:
    results: dict[int, list[dict[str, str]]] = {}
    raw_sha256 = _sha256(raw_path)
    cache_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="bay-outlook-pit-") as temporary:
        directory = Path(temporary)
        for year in years:
            cached = cache_dir / f"pit-{year}.csv"
            cached_metadata = cache_dir / f"pit-{year}.metadata.json"
            if cached.is_file() and cached_metadata.is_file():
                metadata = _read_json(cached_metadata)
                if metadata.get("input_sha256") == raw_sha256:
                    with cached.open(newline="", encoding="utf-8-sig") as stream:
                        results[year] = list(csv.DictReader(stream))
                    continue
            output = directory / f"pit-{year}.csv"
            command = [
                "npx",
                "--yes",
                "xlsx-cli@1.1.3",
                "-s",
                str(year),
                "-o",
                str(output),
                str(raw_path),
            ]
            try:
                subprocess.run(
                    command,
                    check=True,
                    capture_output=True,
                    text=True,
                    timeout=180,
                )
            except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
                raise RuntimeError(
                    "PIT refresh requires Node.js and pinned xlsx-cli@1.1.3 to read HUD's XLSB workbook"
                ) from exc
            with output.open(newline="", encoding="utf-8-sig") as stream:
                results[year] = list(csv.DictReader(stream))
            shutil.copy2(output, cached)
            _write_json(
                cached_metadata,
                {
                    "input_file": str(raw_path),
                    "input_sha256": raw_sha256,
                    "sheet": str(year),
                    "extractor": "xlsx-cli@1.1.3",
                    "output_sha256": _sha256(cached),
                    "row_count": len(results[year]),
                },
            )
    return results


def _pit_hic_observations(
    *, root: Path, data_root: Path, retrieved_at: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    pit_content, pit_snapshot, pit_path = _save_complete_snapshot(
        root=root,
        data_root=data_root,
        source_id="HUD_PIT",
        filename="2007-2025-PIT-Counts-by-CoC.xlsb",
        url=HUD_PIT_URL,
        release="HUD-PIT-2025",
        retrieved_at=retrieved_at,
        landing_url=ACCESS_SOURCE_REGISTRY["HUD_PIT"]["landing_url"],
    )
    if not pit_content.startswith(b"\xd0\xcf\x11\xe0") and not _looks_like_zip(pit_content):
        raise ValueError("HUD PIT response is not an Excel binary workbook")
    sheets = _extract_pit_sheets(
        pit_path,
        PIT_HIC_YEARS,
        cache_dir=data_root / "processed" / "pit",
    )
    observations: list[dict[str, Any]] = []
    count_types: list[dict[str, Any]] = []
    coc_to_fips = {coc: fips for fips, coc in COUNTY_TO_COC.items()}
    pit_metrics = {
        "Overall Homeless": ("pit_overall_homeless", "PIT overall homelessness estimate"),
        "Sheltered Total Homeless": ("pit_sheltered_homeless", "PIT sheltered homelessness estimate"),
        "Unsheltered Homeless": (
            "pit_unsheltered_homeless",
            "PIT unsheltered homelessness estimate",
        ),
    }
    for year, rows in sheets.items():
        found: set[str] = set()
        for row in rows:
            coc = row.get("CoC Number", "")
            if coc not in coc_to_fips:
                continue
            fips = coc_to_fips[coc]
            found.add(fips)
            raw_count_type = row.get("Count Types", "")
            count_type = _normalise_count_type(raw_count_type)
            sheltered_only = count_type == "Sheltered-Only Count"
            count_types.append(
                {
                    "county_fips": fips,
                    "county_name": COUNTY_BY_FIPS[fips],
                    "coc_number": coc,
                    "coc_name": row.get("CoC Name", ""),
                    "period": str(year),
                    "count_type": count_type,
                    "full_sheltered_and_unsheltered_count": not sheltered_only,
                }
            )
            for column, (metric_id, metric_name) in pit_metrics.items():
                observations.append(
                    _observation(
                        domain="homelessness",
                        metric_id=metric_id,
                        metric_name=metric_name,
                        county_fips=fips,
                        period=str(year),
                        period_end=f"{year}-01-31",
                        frequency="annual January point-in-time estimate",
                        value=_number(row.get(column)),
                        unit="people",
                        source_ids="HUD_PIT",
                        source_releases="HUD-PIT-2025-workbook",
                        retrieved_at=retrieved_at,
                        raw_sha256s=pit_snapshot["snapshot_sha256"],
                        notes=(
                            f"HUD count type: {count_type}. "
                            + (
                                "The unsheltered component was not newly counted in this cycle; HUD carries the available component into the overall estimate."
                                if sheltered_only
                                else "Both sheltered and unsheltered components were counted for this cycle."
                            )
                        ),
                        comparability_status=(
                            "sheltered_only_cycle" if sheltered_only else "full_sheltered_and_unsheltered_cycle"
                        ),
                    )
                )
        if found != set(COUNTY_BY_FIPS):
            raise ValueError(f"HUD PIT {year} is missing counties: {sorted(set(COUNTY_BY_FIPS) - found)}")

    hic_content, hic_snapshot, _ = _save_complete_snapshot(
        root=root,
        data_root=data_root,
        source_id="HUD_HIC",
        filename="2007-2025-HIC-Counts-by-CoC.xlsx",
        url=HUD_HIC_URL,
        release="HUD-HIC-2025",
        retrieved_at=retrieved_at,
        landing_url=ACCESS_SOURCE_REGISTRY["HUD_HIC"]["landing_url"],
    )
    workbook = load_workbook(io.BytesIO(hic_content), read_only=True, data_only=True)
    hic_metrics = {
        "Total Year-Round Beds (ES, TH, SH)": (
            "hic_emergency_transitional_safe_haven_beds",
            "Year-round emergency, transitional, and safe-haven beds",
        ),
        "Total Year-Round Beds (PSH)": (
            "hic_permanent_supportive_housing_beds",
            "Year-round permanent supportive housing beds",
        ),
        "Total Year-Round Beds (RRH)": (
            "hic_rapid_rehousing_beds",
            "Year-round rapid rehousing beds",
        ),
    }
    for year in PIT_HIC_YEARS:
        sheet = workbook[str(year)]
        header_row = 1 if year == 2025 else 2
        header = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        index = {value: idx for idx, value in enumerate(header)}
        found: set[str] = set()
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            coc = values[index["CoC Number"]]
            if coc not in coc_to_fips:
                continue
            fips = coc_to_fips[str(coc)]
            found.add(fips)
            for column, (metric_id, metric_name) in hic_metrics.items():
                observations.append(
                    _observation(
                        domain="homelessness",
                        metric_id=metric_id,
                        metric_name=metric_name,
                        county_fips=fips,
                        period=str(year),
                        period_end=f"{year}-01-31",
                        frequency="annual January inventory",
                        value=_number(values[index[column]]),
                        unit="year-round beds",
                        source_ids="HUD_HIC",
                        source_releases="HUD-HIC-2025-workbook",
                        retrieved_at=retrieved_at,
                        raw_sha256s=hic_snapshot["snapshot_sha256"],
                        notes="Inventory capacity is not the same as utilization, exits to housing, or unmet need.",
                    )
                )
        if found != set(COUNTY_BY_FIPS):
            raise ValueError(f"HUD HIC {year} is missing counties: {sorted(set(COUNTY_BY_FIPS) - found)}")
    return observations, [pit_snapshot, hic_snapshot], count_types


def _picture_observations(
    *, root: Path, data_root: Path, retrieved_at: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    observations: list[dict[str, Any]] = []
    snapshots: list[dict[str, Any]] = []
    metric_columns = {
        "total_units": ("hud_assisted_units", "HUD-assisted units under contract", "housing units"),
        "pct_occupied": ("hud_assisted_occupancy_pct", "HUD-assisted units reported occupied", "percent"),
        "people_total": ("hud_assisted_people", "People in HUD-assisted households", "people"),
        "hh_income": (
            "hud_assisted_household_income",
            "Average annual income of HUD-assisted households",
            "current dollars per year",
        ),
        "pct_lt30_median": (
            "hud_assisted_extremely_low_income_pct",
            "HUD-assisted households below 30% of area median income",
            "percent",
        ),
        "months_waiting": (
            "hud_assisted_months_waiting",
            "Average months on waiting list before admission",
            "months",
        ),
    }
    for year in PICTURE_YEARS:
        url = HUD_PICTURE_URL.format(year=year, suffix=HUD_PICTURE_SUFFIX[year])
        content, snapshot, _ = _save_complete_snapshot(
            root=root,
            data_root=data_root,
            source_id="HUD_PICTURE",
            filename=f"COUNTY_{year}_{HUD_PICTURE_SUFFIX[year]}.xlsx",
            url=url,
            release=f"HUD-Picture-{year}",
            retrieved_at=retrieved_at,
            landing_url=ACCESS_SOURCE_REGISTRY["HUD_PICTURE"]["landing_url"],
        )
        snapshots.append(snapshot)
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        index = {str(value).casefold(): idx for idx, value in enumerate(header)}
        found: set[str] = set()
        for values in sheet.iter_rows(min_row=2, values_only=True):
            code = str(values[index["code"]]).zfill(5)
            if code not in COUNTY_BY_FIPS or values[index["program"]] != 1:
                continue
            found.add(code)
            for column, (metric_id, metric_name, unit) in metric_columns.items():
                column_index = index.get(column)
                value = _number(values[column_index]) if column_index is not None else None
                if metric_id == "hud_assisted_months_waiting" and value is not None and value < 0:
                    value = None
                observations.append(
                    _observation(
                        domain="hud_assisted_housing",
                        metric_id=metric_id,
                        metric_name=metric_name,
                        county_fips=code,
                        period=str(year),
                        period_end=f"{year}-12-31",
                        frequency="annual program snapshot",
                        value=value,
                        unit=unit,
                        source_ids="HUD_PICTURE",
                        source_releases=f"HUD-Picture-{year}",
                        retrieved_at=retrieved_at,
                        raw_sha256s=snapshot["snapshot_sha256"],
                        notes="Summary of all programs represented in HUD's Picture of Subsidized Households; it is not a census of every affordable unit.",
                        comparability_status=(
                            "missing_or_not_reported" if value is None else "comparable_within_picture_series"
                        ),
                    )
                )
        if found != set(COUNTY_BY_FIPS):
            raise ValueError(f"HUD Picture {year} is missing counties: {sorted(set(COUNTY_BY_FIPS) - found)}")
    return observations, snapshots


def _fmr_observations(
    *,
    root: Path,
    data_root: Path,
    retrieved_at: str,
    existing_observations: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    content, snapshot, _ = _save_complete_snapshot(
        root=root,
        data_root=data_root,
        source_id="HUD_FMR",
        filename="FY26_FMRs_revised.xlsx",
        url=HUD_FMR_URL,
        release="HUD-FMR-FY2026-revised-effective-2026-05-21",
        retrieved_at=retrieved_at,
        landing_url=ACCESS_SOURCE_REGISTRY["HUD_FMR"]["landing_url"],
    )
    repaired = _repair_xlsx_properties(content)
    workbook = load_workbook(io.BytesIO(repaired), read_only=True, data_only=True)
    sheet = workbook["FY26_FMRs_revised"]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    index = {str(value).casefold(): idx for idx, value in enumerate(header)}
    latest_earnings: dict[str, dict[str, Any]] = {}
    for row in existing_observations:
        if row["metric_id"] != "median_worker_earnings" or row["value"] is None:
            continue
        current = latest_earnings.get(row["county_fips"])
        if current is None or row["period_end"] > current["period_end"]:
            latest_earnings[row["county_fips"]] = row

    observations: list[dict[str, Any]] = []
    found: set[str] = set()
    for values in sheet.iter_rows(min_row=2, values_only=True):
        raw_fips = str(values[index["fips"]] or "")
        fips = raw_fips[:5].zfill(5)
        if fips not in COUNTY_BY_FIPS:
            continue
        found.add(fips)
        fmr_1 = _number(values[index["fmr_1"]])
        fmr_2 = _number(values[index["fmr_2"]])
        common = {
            "county_fips": fips,
            "period": "FY2026",
            "period_end": "2026-09-30",
            "frequency": "federal fiscal year",
            "retrieved_at": retrieved_at,
        }
        area_name = str(values[index["hud_area_name"]])
        for bedroom, fmr in ((1, fmr_1), (2, fmr_2)):
            required_income = fmr * 12 / 0.30 if fmr is not None else None
            observations.extend(
                [
                    _observation(
                        domain="worker_housing_access",
                        metric_id=f"fmr_{bedroom}br",
                        metric_name=f"HUD Fair Market Rent — {bedroom}-bedroom",
                        value=fmr,
                        unit="current dollars per month",
                        source_ids="HUD_FMR",
                        source_releases=snapshot["source_release"],
                        raw_sha256s=snapshot["snapshot_sha256"],
                        notes=f"HUD FMR area: {area_name}. FMR is a rent standard, not the median asking rent for this county alone.",
                        comparability_status="hud_area_value_assigned_to_county",
                        **common,
                    ),
                    _observation(
                        domain="worker_housing_access",
                        metric_id=f"income_required_{bedroom}br_fmr",
                        metric_name=f"Annual income required for {bedroom}-bedroom FMR at 30% of income",
                        value=required_income,
                        unit="current dollars per year",
                        source_ids="HUD_FMR",
                        source_releases=snapshot["source_release"],
                        raw_sha256s=snapshot["snapshot_sha256"],
                        calculation=f"12 × {bedroom}-bedroom monthly FMR divided by 0.30",
                        notes="Descriptive affordability benchmark; it does not account for household size, taxes, benefits, or work hours.",
                        comparability_status="hud_area_value_assigned_to_county",
                        **common,
                    ),
                ]
            )
            earnings = latest_earnings.get(fips)
            if earnings is None:
                continue
            combined_sources = "CENSUS_ACS5_BULK;HUD_FMR"
            combined_releases = f"{earnings['source_releases']};{snapshot['source_release']}"
            combined_hashes = f"{earnings['raw_sha256s']};{snapshot['snapshot_sha256']}"
            coverage = _ratio(earnings["value"], required_income)
            affordable_monthly = earnings["value"] * 0.30 / 12 if earnings["value"] is not None else None
            gap = fmr - affordable_monthly if fmr is not None and affordable_monthly is not None else None
            observations.extend(
                [
                    _observation(
                        domain="worker_housing_access",
                        metric_id=f"median_worker_earnings_coverage_{bedroom}br_pct",
                        metric_name=f"Median worker earnings as share of income required for {bedroom}-bedroom FMR",
                        value=coverage,
                        unit="percent",
                        source_ids=combined_sources,
                        source_releases=combined_releases,
                        raw_sha256s=combined_hashes,
                        calculation=f"ACS 2024 median earnings divided by annual income required for FY2026 {bedroom}-bedroom FMR",
                        notes="Comparison joins overlapping ACS 2019–2024 earnings with an FY2026 rent standard; it is not occupation-specific.",
                        comparability_status="acs_2024_compared_with_fy2026_fmr",
                        **common,
                    ),
                    _observation(
                        domain="worker_housing_access",
                        metric_id=f"monthly_affordability_gap_{bedroom}br",
                        metric_name=f"Monthly gap between {bedroom}-bedroom FMR and 30% of median worker earnings",
                        value=gap,
                        unit="current dollars per month",
                        source_ids=combined_sources,
                        source_releases=combined_releases,
                        raw_sha256s=combined_hashes,
                        calculation=f"FY2026 {bedroom}-bedroom FMR minus 30% of ACS 2024 median annual earnings divided by 12",
                        notes="Positive values indicate the rent standard exceeds 30% of median individual earnings; no work-hour assumption is used.",
                        comparability_status="acs_2024_compared_with_fy2026_fmr",
                        **common,
                    ),
                ]
            )
    if found != set(COUNTY_BY_FIPS):
        raise ValueError(f"HUD FMR is missing counties: {sorted(set(COUNTY_BY_FIPS) - found)}")
    return observations, [snapshot]


def _latest_snapshot_rows(observations: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest: dict[tuple[str, str], dict[str, Any]] = {}
    for row in observations:
        key = (row["metric_id"], row["county_fips"])
        current = latest.get(key)
        if current is None or (row["period_end"], row["period"]) > (
            current["period_end"],
            current["period"],
        ):
            latest[key] = row
    return sorted(latest.values(), key=lambda row: (row["domain"], row["metric_id"], row["county_name"]))


def _income_progress(observations: list[dict[str, Any]]) -> list[dict[str, Any]]:
    categories = {
        "vli": ("hcd_apr_vli_permitted_units", "rhna_target_vli_units", "Very low income"),
        "li": ("hcd_apr_li_permitted_units", "rhna_target_li_units", "Low income"),
        "mod": ("hcd_apr_mod_permitted_units", "rhna_target_mod_units", "Moderate income"),
        "above": (
            "hcd_apr_above_permitted_units",
            "rhna_target_above_mod_units",
            "Above moderate income",
        ),
    }
    rows: list[dict[str, Any]] = []
    for fips, county in COUNTY_BY_FIPS.items():
        for category, (permit_metric, target_metric, label) in categories.items():
            permitted = sum(
                float(row["value"])
                for row in observations
                if row["county_fips"] == fips
                and row["metric_id"] == permit_metric
                and row["value"] is not None
            )
            target_row = next(
                row
                for row in observations
                if row["county_fips"] == fips and row["metric_id"] == target_metric
            )
            target = target_row["value"]
            rows.append(
                {
                    "county_fips": fips,
                    "county_name": county,
                    "category": category,
                    "category_label": label,
                    "permitted_2023_2025": permitted,
                    "sixth_cycle_target": target,
                    "progress_pct": _ratio(permitted, target),
                    "permit_source": "CA_HCD_APR",
                    "target_source": "CA_HCD_RHNA",
                    "calculation": "sum of 2023–2025 permits divided by sixth-cycle allocation",
                }
            )
    return rows


def _source_freshness(
    observations: list[dict[str, Any]], snapshots: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    latest_period: dict[str, str] = {}
    latest_retrieval: dict[str, str] = {}
    snapshot_counts: dict[str, int] = defaultdict(int)
    for row in observations:
        for source_id in row["source_ids"].split(";"):
            latest_period[source_id] = max(latest_period.get(source_id, ""), row["period_end"])
    for row in snapshots:
        snapshot_counts[row["source_id"]] += 1
        latest_retrieval[row["source_id"]] = max(
            latest_retrieval.get(row["source_id"], ""), row["retrieved_at"]
        )
    return [
        {
            "source_id": source_id,
            "publisher": source["publisher"],
            "dataset": source["title"],
            "source_tier": source["source_tier"],
            "source_class": source["source_class"],
            "frequency": source["frequency"],
            "latest_period_end": latest_period.get(source_id, ""),
            "retrieved_at": latest_retrieval.get(source_id, ""),
            "snapshot_count": snapshot_counts[source_id],
            "status": "current for documented release cycle",
            "landing_url": source["landing_url"],
        }
        for source_id, source in ACCESS_SOURCE_REGISTRY.items()
    ]


def _quality_checks(
    observations: list[dict[str, Any]],
    snapshots: list[dict[str, Any]],
    count_types: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []

    def add(name: str, passed: bool, detail: Any) -> None:
        checks.append({"check_name": name, "passed": int(bool(passed)), "detail": detail})

    domains = {row["domain"] for row in observations}
    add("six_domain_scope", domains == REQUIRED_ACCESS_DOMAINS, sorted(domains))
    coverage: dict[str, list[str]] = {}
    for domain, metric_id in CRITICAL_ACCESS_METRICS.items():
        coverage[domain] = sorted(
            {row["county_fips"] for row in observations if row["metric_id"] == metric_id}
        )
    add(
        "nine_county_critical_coverage",
        all(set(counties) == set(COUNTY_BY_FIPS) for counties in coverage.values()),
        coverage,
    )

    history = defaultdict(set)
    for row in observations:
        if row["value"] is not None:
            history[(row["metric_id"], row["county_fips"])].add(row["period"])
    requirements = {
        "unlawful_detainer_filings": 11,
        "renter_overcrowding_pct": 4,
        "pit_overall_homeless": 4,
        "hic_permanent_supportive_housing_beds": 4,
        "hud_assisted_units": 4,
        "hcd_apr_vli_permitted_units": 3,
        "median_worker_earnings": 4,
        "fmr_1br": 1,
    }
    missing_history = {
        f"{metric}:{fips}": len(history[(metric, fips)])
        for metric, minimum in requirements.items()
        for fips in COUNTY_BY_FIPS
        if len(history[(metric, fips)]) < minimum
    }
    add("documented_history_minimums", not missing_history, missing_history)

    keys: dict[tuple[str, str, str], int] = defaultdict(int)
    for row in observations:
        keys[(row["metric_id"], row["county_fips"], row["period"])] += 1
    duplicates = [list(key) for key, count in keys.items() if count > 1]
    add("natural_key_uniqueness", not duplicates, duplicates[:20])

    percent_metrics = {
        row["metric_id"] for row in observations if row["unit"] == "percent"
    }
    invalid_percent = [
        {"metric": row["metric_id"], "county": row["county_fips"], "value": row["value"]}
        for row in observations
        if row["metric_id"] in percent_metrics
        and row["value"] is not None
        and not (0 <= float(row["value"]) <= 100)
    ]
    add("percentage_ranges", not invalid_percent, invalid_percent[:20])

    snapshot_sources = {row["source_id"] for row in snapshots}
    observation_sources = {
        source_id for row in observations for source_id in row["source_ids"].split(";")
    }
    add(
        "complete_source_lineage",
        observation_sources == set(ACCESS_SOURCE_REGISTRY) and observation_sources <= snapshot_sources,
        {"observations": sorted(observation_sources), "snapshots": sorted(snapshot_sources)},
    )

    sheltered_only_2025 = sorted(
        row["county_fips"]
        for row in count_types
        if row["period"] == "2025" and not row["full_sheltered_and_unsheltered_count"]
    )
    add(
        "pit_count_type_disclosure",
        len(count_types) == 36
        and sheltered_only_2025 == ["06001", "06041", "06075", "06081", "06095"],
        {"records": len(count_types), "sheltered_only_2025": sheltered_only_2025},
    )

    opaque_tokens = ("risk_score", "risk_index", "displacement_score", "displacement_index")
    opaque_metrics = [
        row["metric_id"]
        for row in observations
        if any(token in row["metric_id"] for token in opaque_tokens)
    ]
    add("transparent_pressure_components_only", not opaque_metrics, sorted(set(opaque_metrics)))

    calculation_errors: list[dict[str, Any]] = []
    lookup = {(row["metric_id"], row["county_fips"], row["period"]): row for row in observations}
    for fips in COUNTY_BY_FIPS:
        for year in (2021, 2022, 2023):
            period = f"FY{year}-{str(year + 1)[-2:]}"
            filing = lookup.get(("unlawful_detainer_filings", fips, period))
            renter = lookup.get(("renter_households", fips, str(year)))
            rate = lookup.get(("unlawful_detainer_filings_per_1000_renter_households", fips, period))
            expected = _ratio(filing["value"], renter["value"], multiplier=1000.0) if filing and renter else None
            if expected is None or rate is None or abs(float(rate["value"]) - expected) > 1e-9:
                calculation_errors.append({"metric": "eviction_rate", "county": fips, "period": period})
    add("derived_calculation_reproducibility", not calculation_errors, calculation_errors)
    return checks


def _build_database(
    path: Path,
    observations: list[dict[str, Any]],
    snapshots: list[dict[str, Any]],
    checks: list[dict[str, Any]],
    built_at: str,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix="phase14-access-", suffix=".sqlite", dir=path.parent, delete=False
    ) as temporary_handle:
        temporary = Path(temporary_handle.name)
    try:
        with sqlite3.connect(temporary) as connection:
            connection.executescript(
                """
                PRAGMA foreign_keys = ON;
                CREATE TABLE build_metadata (
                    metadata_key TEXT PRIMARY KEY,
                    metadata_value TEXT NOT NULL
                );
                CREATE TABLE source_registry (
                    source_id TEXT PRIMARY KEY,
                    publisher TEXT NOT NULL,
                    title TEXT NOT NULL,
                    source_tier INTEGER NOT NULL,
                    source_class TEXT NOT NULL,
                    frequency TEXT NOT NULL,
                    geography_basis TEXT NOT NULL,
                    landing_url TEXT NOT NULL
                );
                CREATE TABLE source_snapshot (
                    snapshot_id INTEGER PRIMARY KEY,
                    source_id TEXT NOT NULL REFERENCES source_registry(source_id),
                    source_url TEXT NOT NULL,
                    source_release TEXT NOT NULL,
                    retrieved_at TEXT NOT NULL,
                    snapshot_sha256 TEXT NOT NULL,
                    upstream_sha256 TEXT NOT NULL,
                    byte_count INTEGER,
                    upstream_byte_count INTEGER,
                    snapshot_kind TEXT NOT NULL,
                    raw_path TEXT NOT NULL UNIQUE,
                    metadata_path TEXT NOT NULL UNIQUE
                );
                CREATE TABLE access_observation (
                    observation_id INTEGER PRIMARY KEY,
                    domain TEXT NOT NULL,
                    metric_id TEXT NOT NULL,
                    metric_name TEXT NOT NULL,
                    county_fips TEXT NOT NULL,
                    county_name TEXT NOT NULL,
                    period TEXT NOT NULL,
                    period_end TEXT NOT NULL,
                    frequency TEXT NOT NULL,
                    value REAL,
                    unit TEXT NOT NULL,
                    margin_of_error REAL,
                    retrieved_at TEXT NOT NULL,
                    calculation TEXT NOT NULL,
                    notes TEXT NOT NULL,
                    comparability_status TEXT NOT NULL,
                    UNIQUE(metric_id, county_fips, period)
                );
                CREATE TABLE observation_source (
                    observation_id INTEGER NOT NULL REFERENCES access_observation(observation_id),
                    source_id TEXT NOT NULL REFERENCES source_registry(source_id),
                    source_release TEXT NOT NULL,
                    raw_sha256 TEXT NOT NULL,
                    source_order INTEGER NOT NULL,
                    PRIMARY KEY(observation_id, source_id)
                );
                CREATE TABLE quality_check (
                    check_name TEXT PRIMARY KEY,
                    passed INTEGER NOT NULL CHECK(passed IN (0, 1)),
                    detail TEXT NOT NULL
                );
                CREATE INDEX idx_access_metric_county_period
                    ON access_observation(metric_id, county_fips, period_end);
                CREATE INDEX idx_access_domain_period
                    ON access_observation(domain, period_end);
                """
            )
            connection.executemany(
                "INSERT INTO build_metadata(metadata_key, metadata_value) VALUES (?, ?)",
                [
                    ("project", "The Bay Outlook"),
                    ("phase", "14"),
                    ("version", PHASE14_ACCESS_VERSION),
                    ("product", "Housing Access & Displacement"),
                    ("built_at", built_at),
                    ("publication_authority", "named human only"),
                    ("composite_displacement_score", "not calculated"),
                ],
            )
            connection.executemany(
                """INSERT INTO source_registry(
                    source_id, publisher, title, source_tier, source_class,
                    frequency, geography_basis, landing_url
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                [
                    (
                        source_id,
                        row["publisher"],
                        row["title"],
                        row["source_tier"],
                        row["source_class"],
                        row["frequency"],
                        row["geography_basis"],
                        row["landing_url"],
                    )
                    for source_id, row in ACCESS_SOURCE_REGISTRY.items()
                ],
            )
            connection.executemany(
                """INSERT INTO source_snapshot(
                    source_id, source_url, source_release, retrieved_at,
                    snapshot_sha256, upstream_sha256, byte_count, upstream_byte_count,
                    snapshot_kind, raw_path, metadata_path
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                [
                    tuple(
                        row[key]
                        for key in (
                            "source_id",
                            "source_url",
                            "source_release",
                            "retrieved_at",
                            "snapshot_sha256",
                            "upstream_sha256",
                            "byte_count",
                            "upstream_byte_count",
                            "snapshot_kind",
                            "raw_path",
                            "metadata_path",
                        )
                    )
                    for row in snapshots
                ],
            )
            for row in observations:
                cursor = connection.execute(
                    """INSERT INTO access_observation(
                        domain, metric_id, metric_name, county_fips, county_name,
                        period, period_end, frequency, value, unit, margin_of_error,
                        retrieved_at, calculation, notes, comparability_status
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    tuple(
                        row[key]
                        for key in (
                            "domain",
                            "metric_id",
                            "metric_name",
                            "county_fips",
                            "county_name",
                            "period",
                            "period_end",
                            "frequency",
                            "value",
                            "unit",
                            "margin_of_error",
                            "retrieved_at",
                            "calculation",
                            "notes",
                            "comparability_status",
                        )
                    ),
                )
                source_ids = row["source_ids"].split(";")
                releases = row["source_releases"].split(";")
                hashes = row["raw_sha256s"].split(";")
                if not (len(source_ids) == len(releases) == len(hashes)):
                    raise ValueError(f"misaligned source lineage for {row['metric_id']}")
                connection.executemany(
                    """INSERT INTO observation_source(
                        observation_id, source_id, source_release, raw_sha256, source_order
                    ) VALUES (?, ?, ?, ?, ?)""",
                    [
                        (cursor.lastrowid, source_id, releases[index], hashes[index], index)
                        for index, source_id in enumerate(source_ids)
                    ],
                )
            connection.executemany(
                "INSERT INTO quality_check(check_name, passed, detail) VALUES (?, ?, ?)",
                [
                    (row["check_name"], row["passed"], json.dumps(row["detail"], sort_keys=True))
                    for row in checks
                ],
            )
            connection.commit()
            integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
            foreign_keys = connection.execute("PRAGMA foreign_key_check").fetchall()
            if integrity != "ok" or foreign_keys:
                raise ValueError(
                    f"Version 1.2 database integrity failed: integrity={integrity}, foreign_keys={foreign_keys}"
                )
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _public_payload(
    *,
    observations: list[dict[str, Any]],
    count_types: list[dict[str, Any]],
    income_progress: list[dict[str, Any]],
    freshness: list[dict[str, Any]],
    built_at: str,
) -> dict[str, Any]:
    metric_catalog: dict[str, dict[str, Any]] = {}
    for row in observations:
        metric_catalog.setdefault(
            row["metric_id"],
            {
                "metricId": row["metric_id"],
                "metricName": row["metric_name"],
                "domain": row["domain"],
                "unit": row["unit"],
                "frequency": row["frequency"],
                "sourceIds": row["source_ids"].split(";"),
                "calculation": row["calculation"],
                "notes": row["notes"],
            },
        )
    public_observations = [
        {
            "domain": row["domain"],
            "metricId": row["metric_id"],
            "countyFips": row["county_fips"],
            "countyName": row["county_name"],
            "period": row["period"],
            "periodEnd": row["period_end"],
            "value": row["value"],
            "unit": row["unit"],
            "marginOfError": row["margin_of_error"],
            "sourceIds": row["source_ids"].split(";"),
            "sourceReleases": row["source_releases"].split(";"),
            "comparabilityStatus": row["comparability_status"],
        }
        for row in observations
    ]
    return {
        "project": "The Bay Outlook",
        "product": "Housing Access & Displacement",
        "version": PHASE14_ACCESS_VERSION,
        "builtAt": built_at,
        "evidenceAsOf": max(
            row["period_end"]
            for row in observations
            if row["period_end"] <= built_at[:10]
        ),
        "countyCount": len(COUNTY_BY_FIPS),
        "domainCount": len(REQUIRED_ACCESS_DOMAINS),
        "metricCount": len(metric_catalog),
        "observationCount": len(observations),
        "counties": [
            {
                "countyFips": fips,
                "countyName": name,
                "cocNumber": COUNTY_TO_COC[fips],
            }
            for fips, name in COUNTY_BY_FIPS.items()
        ],
        "domains": sorted(REQUIRED_ACCESS_DOMAINS),
        "metricCatalog": sorted(metric_catalog.values(), key=lambda row: row["metricId"]),
        "observations": public_observations,
        "pitCountTypes": [
            {
                "countyFips": row["county_fips"],
                "countyName": row["county_name"],
                "cocNumber": row["coc_number"],
                "cocName": row["coc_name"],
                "period": row["period"],
                "countType": row["count_type"],
                "fullShelteredAndUnshelteredCount": row[
                    "full_sheltered_and_unsheltered_count"
                ],
            }
            for row in sorted(count_types, key=lambda item: (item["period"], item["county_name"]))
        ],
        "incomeProgress": [
            {
                "countyFips": row["county_fips"],
                "countyName": row["county_name"],
                "category": row["category"],
                "categoryLabel": row["category_label"],
                "permitted2023To2025": row["permitted_2023_2025"],
                "sixthCycleTarget": row["sixth_cycle_target"],
                "progressPct": row["progress_pct"],
                "permitSource": row["permit_source"],
                "targetSource": row["target_source"],
            }
            for row in income_progress
        ],
        "sources": freshness,
        "pressureFramework": {
            "compositeScore": None,
            "rankingProduced": False,
            "components": [
                "limited civil unlawful-detainer filings",
                "severe renter cost burden",
                "renter overcrowding",
                "median gross rent",
            ],
            "interpretation": "Read components together; they are not combined into an opaque displacement-risk score.",
        },
        "limitations": [
            "Court filings and dispositions are legal-process counts, not executed evictions; unlimited cases and informal displacement are omitted.",
            "PIT estimates are one-night counts. A sheltered-only cycle can reuse an unsheltered component from an earlier count, so year-to-year changes require count-type context.",
            "HIC beds measure inventory, not utilization, exits to housing, or unmet need.",
            "HUD Picture covers programs represented in that administrative extract and is not a census of all income-restricted or subsidized housing.",
            "ACS five-year estimates overlap across vintages; adjacent releases are not independent annual samples.",
            "Worker access compares county median individual earnings with an FMR-area rent standard; it is not occupation- or household-specific.",
            "HCD permit counts are jurisdiction-reported authorizations and do not prove completion or occupancy.",
        ],
        "publicationBoundary": {
            "dataStatus": "public descriptive evidence",
            "automatedNarrative": False,
            "newAnalysisRequiresHumanApproval": True,
            "phase10ReportStatus": "human_approval_hold",
        },
    }


def build_phase14_access(
    *,
    root: Path | None = None,
    built_at: str | None = None,
    refresh: bool = True,
) -> dict[str, Any]:
    root = (root or PROJECT_ROOT).resolve()
    built_at = built_at or _utc_now()
    output = root / "data" / "phase14" / "access"
    exports = output / "exports"
    if not refresh:
        if not (exports / "access_time_series.csv").is_file():
            raise FileNotFoundError("Version 1.2 offline build requires existing access exports")
        return _build_manifest(root=root, built_at=built_at)

    inherited_verification = verify_phase14(root=root)
    if not inherited_verification["complete"]:
        raise ValueError(
            f"Version 1.2 requires a verified Version 1.1 baseline: {inherited_verification['failed']}"
        )

    observations: list[dict[str, Any]] = []
    snapshots: list[dict[str, Any]] = []

    inherited_observations, inherited_snapshots, latest_inherited_snapshots = (
        _inherited_v1_1_observations(root=root, retrieved_at=built_at)
    )
    observations.extend(inherited_observations)
    snapshots.extend(inherited_snapshots)

    acs_observations, acs_snapshots, renter_households = _acs_access_observations(
        root=root,
        data_root=output,
        retrieved_at=built_at,
    )
    observations.extend(acs_observations)
    snapshots.extend(acs_snapshots)

    court_observations, court_snapshots = _court_observations(
        root=root,
        data_root=output,
        retrieved_at=built_at,
        renter_households=renter_households,
    )
    observations.extend(court_observations)
    snapshots.extend(court_snapshots)

    observations.extend(
        _income_target_observations(
            root=root,
            retrieved_at=built_at,
            latest_snapshots=latest_inherited_snapshots,
        )
    )

    homeless_observations, homeless_snapshots, count_types = _pit_hic_observations(
        root=root,
        data_root=output,
        retrieved_at=built_at,
    )
    observations.extend(homeless_observations)
    snapshots.extend(homeless_snapshots)

    picture_observations, picture_snapshots = _picture_observations(
        root=root,
        data_root=output,
        retrieved_at=built_at,
    )
    observations.extend(picture_observations)
    snapshots.extend(picture_snapshots)

    fmr_observations, fmr_snapshots = _fmr_observations(
        root=root,
        data_root=output,
        retrieved_at=built_at,
        existing_observations=observations,
    )
    observations.extend(fmr_observations)
    snapshots.extend(fmr_snapshots)

    observations.sort(
        key=lambda row: (row["domain"], row["metric_id"], row["county_fips"], row["period_end"])
    )
    unique_snapshots: dict[str, dict[str, Any]] = {}
    for row in snapshots:
        unique_snapshots[row["raw_path"]] = row
    snapshots = sorted(
        unique_snapshots.values(),
        key=lambda row: (row["source_id"], row["source_release"], row["raw_path"]),
    )

    checks = _quality_checks(observations, snapshots, count_types)
    failures = [row for row in checks if not row["passed"]]
    if failures:
        raise ValueError(f"Version 1.2 quality checks failed: {failures}")

    latest_rows = _latest_snapshot_rows(observations)
    income_progress = _income_progress(observations)
    freshness = _source_freshness(observations, snapshots)
    _write_csv(exports / "access_time_series.csv", observations, OBSERVATION_FIELDS)
    _write_csv(exports / "access_snapshot.csv", latest_rows, OBSERVATION_FIELDS)
    _write_csv(exports / "income_progress.csv", income_progress)
    _write_csv(exports / "pit_count_types.csv", count_types)
    _write_csv(exports / "source_freshness.csv", freshness)
    _write_csv(exports / "source_snapshots.csv", snapshots)
    _write_csv(exports / "quality_checks.csv", checks)

    database_path = output / "housing_access.sqlite"
    _build_database(database_path, observations, snapshots, checks, built_at)
    payload = _public_payload(
        observations=observations,
        count_types=count_types,
        income_progress=income_progress,
        freshness=freshness,
        built_at=built_at,
    )
    payload_path = output / "public" / "housing-access-data.json"
    _write_json(payload_path, payload)

    manifest = _build_manifest(root=root, built_at=built_at)
    manifest["build"] = {
        "observations": len(observations),
        "metrics": len({row["metric_id"] for row in observations}),
        "snapshots": len(snapshots),
        "quality_checks": len(checks),
        "pit_count_type_records": len(count_types),
    }
    return manifest


def _build_manifest(*, root: Path, built_at: str) -> dict[str, Any]:
    output = root / "data" / "phase14" / "access"
    manifest_path = output / "phase14_v1_2_manifest.json"
    with (output / "exports" / "access_time_series.csv").open(
        newline="", encoding="utf-8"
    ) as stream:
        observations = list(csv.DictReader(stream))
    with (output / "exports" / "source_snapshots.csv").open(
        newline="", encoding="utf-8"
    ) as stream:
        snapshots = list(csv.DictReader(stream))
    with (output / "exports" / "quality_checks.csv").open(
        newline="", encoding="utf-8"
    ) as stream:
        quality = list(csv.DictReader(stream))
    missing = [relative for relative in REQUIRED_ACCESS_FILES if not (root / relative).is_file()]
    files = {
        relative: _sha256(root / relative)
        for relative in REQUIRED_ACCESS_FILES
        if (root / relative).is_file()
    }
    manifest = {
        "project": "The Bay Outlook",
        "phase": 14,
        "product": "Housing Access & Displacement",
        "product_version": PHASE14_ACCESS_VERSION,
        "built_at": built_at,
        "completion_state": "complete" if not missing else "incomplete",
        "scope": sorted(REQUIRED_ACCESS_DOMAINS),
        "counts": {
            "counties": len({row["county_fips"] for row in observations}),
            "domains": len({row["domain"] for row in observations}),
            "metrics": len({row["metric_id"] for row in observations}),
            "observations": len(observations),
            "sources": len(
                {
                    source_id
                    for row in observations
                    for source_id in row["source_ids"].split(";")
                }
            ),
            "source_snapshots": len(snapshots),
            "quality_checks": len(quality),
        },
        "design_boundary": {
            "composite_displacement_score": False,
            "county_ranking": False,
            "transparent_components_only": True,
        },
        "phase10_hold": {
            "status": "human_approval_hold",
            "human_signoff": 0,
            "approved_by": None,
            "published_at": None,
        },
        "missing_required_files": missing,
        "files": files,
    }
    _write_json(manifest_path, manifest)
    verification = verify_phase14_access(manifest_path=manifest_path, root=root)
    manifest["verification"] = {
        "complete": verification["complete"],
        "passing": verification["passing"],
        "total": verification["total"],
        "failed": verification["failed"],
    }
    _write_json(manifest_path, manifest)
    return {"manifest": str(manifest_path), "verification": verification}


def _check(name: str, passed: bool, detail: Any) -> dict[str, Any]:
    return {"check": name, "passed": bool(passed), "detail": detail}


def verify_phase14_access(
    manifest_path: Path | None = None,
    *,
    root: Path | None = None,
) -> dict[str, Any]:
    root = (root or PROJECT_ROOT).resolve()
    manifest_path = manifest_path or (
        root / "data" / "phase14" / "access" / "phase14_v1_2_manifest.json"
    )
    manifest = _read_json(manifest_path)
    checks: list[dict[str, Any]] = []
    checks.append(
        _check(
            "manifest_identity",
            manifest.get("project") == "The Bay Outlook"
            and manifest.get("phase") == 14
            and manifest.get("product") == "Housing Access & Displacement"
            and manifest.get("product_version") == PHASE14_ACCESS_VERSION,
            {
                "project": manifest.get("project"),
                "phase": manifest.get("phase"),
                "product": manifest.get("product"),
                "version": manifest.get("product_version"),
            },
        )
    )

    inherited = verify_phase14(root=root)
    checks.append(
        _check(
            "verified_version_1_1_baseline",
            inherited["complete"],
            {"passing": inherited["passing"], "total": inherited["total"], "failed": inherited["failed"]},
        )
    )

    missing = [relative for relative in REQUIRED_ACCESS_FILES if not (root / relative).is_file()]
    checks.append(_check("required_version_1_2_files", not missing, missing))

    output = root / "data" / "phase14" / "access"
    payload_path = output / "public" / "housing-access-data.json"
    payload = _read_json(payload_path) if payload_path.is_file() else {}
    checks.append(
        _check(
            "six_access_domains",
            set(payload.get("domains", [])) == REQUIRED_ACCESS_DOMAINS,
            payload.get("domains", []),
        )
    )
    checks.append(
        _check(
            "nine_county_coverage",
            payload.get("countyCount") == 9
            and {row.get("countyFips") for row in payload.get("counties", [])}
            == set(COUNTY_BY_FIPS),
            payload.get("counties", []),
        )
    )

    observation_rows = payload.get("observations", [])
    history: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in observation_rows:
        if row.get("value") is not None:
            history[(row["metricId"], row["countyFips"])].add(row["period"])
    requirements = {
        "unlawful_detainer_filings": 11,
        "renter_overcrowding_pct": 4,
        "pit_overall_homeless": 4,
        "hud_assisted_units": 4,
        "hcd_apr_vli_permitted_units": 3,
        "median_worker_earnings": 4,
    }
    history_errors = {
        f"{metric}:{fips}": len(history[(metric, fips)])
        for metric, minimum in requirements.items()
        for fips in COUNTY_BY_FIPS
        if len(history[(metric, fips)]) < minimum
    }
    checks.append(_check("multi_year_history", not history_errors, history_errors))

    quality_path = output / "exports" / "quality_checks.csv"
    with quality_path.open(newline="", encoding="utf-8") as stream:
        quality = list(csv.DictReader(stream))
    failed_quality = [row["check_name"] for row in quality if row["passed"] != "1"]
    checks.append(_check("all_build_quality_checks", not failed_quality and len(quality) == 9, failed_quality))

    source_ids = {
        source_id
        for row in observation_rows
        for source_id in row.get("sourceIds", [])
    }
    checks.append(
        _check(
            "documented_source_lineage",
            source_ids == set(ACCESS_SOURCE_REGISTRY)
            and len(payload.get("sources", [])) == len(ACCESS_SOURCE_REGISTRY),
            sorted(source_ids),
        )
    )

    raw_errors: list[dict[str, str]] = []
    snapshots_path = output / "exports" / "source_snapshots.csv"
    with snapshots_path.open(newline="", encoding="utf-8") as stream:
        snapshot_rows = list(csv.DictReader(stream))
    for row in snapshot_rows:
        raw_path = root / row["raw_path"]
        metadata_path = root / row["metadata_path"]
        if not raw_path.is_file() or not metadata_path.is_file():
            raw_errors.append({"file": row["raw_path"], "reason": "missing"})
            continue
        if _sha256(raw_path) != row["snapshot_sha256"]:
            raw_errors.append({"file": row["raw_path"], "reason": "hash_mismatch"})
        metadata = _read_json(metadata_path)
        if metadata.get("snapshot_kind") == "source_slice" and not metadata.get("upstream_sha256"):
            raw_errors.append({"file": row["metadata_path"], "reason": "missing_upstream_digest"})
    checks.append(
        _check(
            "immutable_raw_evidence",
            not raw_errors and len(snapshot_rows) >= 20,
            {"snapshots": len(snapshot_rows), "errors": raw_errors},
        )
    )

    database = output / "housing_access.sqlite"
    database_detail: dict[str, Any] = {"exists": database.is_file()}
    database_passed = False
    if database.is_file():
        with sqlite3.connect(f"file:{database.resolve()}?mode=ro", uri=True) as connection:
            integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
            foreign_keys = connection.execute("PRAGMA foreign_key_check").fetchall()
            failed = connection.execute("SELECT COUNT(*) FROM quality_check WHERE passed = 0").fetchone()[0]
            count = connection.execute("SELECT COUNT(*) FROM access_observation").fetchone()[0]
            source_links = connection.execute("SELECT COUNT(*) FROM observation_source").fetchone()[0]
        database_detail.update(
            {
                "integrity": integrity,
                "foreign_keys": len(foreign_keys),
                "failed_quality": failed,
                "observations": count,
                "source_links": source_links,
            }
        )
        database_passed = (
            integrity == "ok"
            and not foreign_keys
            and failed == 0
            and count == payload.get("observationCount")
            and source_links >= count
        )
    checks.append(_check("database_integrity", database_passed, database_detail))

    count_type_rows = payload.get("pitCountTypes", [])
    sheltered_only_2025 = sorted(
        row["countyFips"]
        for row in count_type_rows
        if row["period"] == "2025" and not row["fullShelteredAndUnshelteredCount"]
    )
    checks.append(
        _check(
            "pit_comparability_disclosure",
            len(count_type_rows) == 36
            and sheltered_only_2025 == ["06001", "06041", "06075", "06081", "06095"],
            {"records": len(count_type_rows), "sheltered_only_2025": sheltered_only_2025},
        )
    )

    pressure = payload.get("pressureFramework", {})
    checks.append(
        _check(
            "no_opaque_displacement_score",
            pressure.get("compositeScore") is None
            and pressure.get("rankingProduced") is False
            and len(pressure.get("components", [])) >= 4,
            pressure,
        )
    )

    income_errors = [
        row
        for row in payload.get("incomeProgress", [])
        if row["sixthCycleTarget"] in (None, 0)
        or abs(
            float(row["progressPct"])
            - float(row["permitted2023To2025"]) / float(row["sixthCycleTarget"]) * 100
        )
        > 1e-9
    ]
    checks.append(
        _check(
            "income_progress_reproducibility",
            not income_errors and len(payload.get("incomeProgress", [])) == 36,
            income_errors[:10],
        )
    )

    workflow = root / ".github" / "workflows" / "housing-access-update.yml"
    workflow_text = workflow.read_text(encoding="utf-8") if workflow.is_file() else ""
    workflow_tokens = (
        "workflow_dispatch",
        "schedule:",
        "verify-phase14-access",
        "human-review-required",
        "contents: read",
        "No public deployment",
    )
    missing_tokens = [token for token in workflow_tokens if token not in workflow_text]
    checks.append(_check("responsible_update_workflow", not missing_tokens, missing_tokens))

    report = manifest.get("phase10_hold", {})
    hold_passed = (
        report.get("status") == "human_approval_hold"
        and report.get("human_signoff") == 0
        and report.get("approved_by") is None
        and report.get("published_at") is None
    )
    checks.append(_check("phase10_human_approval_hold", hold_passed, report))

    hash_errors: list[dict[str, str]] = []
    for relative, expected in manifest.get("files", {}).items():
        path = root / relative
        if not path.is_file():
            hash_errors.append({"file": relative, "reason": "missing"})
        elif _sha256(path) != expected:
            hash_errors.append({"file": relative, "reason": "hash_mismatch"})
    checks.append(_check("manifest_file_hashes", not hash_errors, hash_errors))

    failed = [item["check"] for item in checks if not item["passed"]]
    return {
        "phase": 14,
        "product": "Housing Access & Displacement",
        "product_version": PHASE14_ACCESS_VERSION,
        "complete": not failed,
        "passing": len(checks) - len(failed),
        "total": len(checks),
        "failed": failed,
        "checks": checks,
    }
